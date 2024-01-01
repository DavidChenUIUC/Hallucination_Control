import json, time, openai
import requests
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from tenacity import retry, wait_random_exponential, stop_after_attempt
from pydantic import BaseModel, confloat
from typing import Literal
import warnings
from tqdm import tqdm
from pprint import pprint
from datetime import datetime
import pandas as pd  # Ensure 'pandas' and 'openpyxl' are installed.
from collections import defaultdict
warnings.filterwarnings("ignore")


class HallucinationDetection(BaseModel):
    hallucination_type: Literal['intrinsic', 'extrinsic', 'both', 'NULL']
    hallucination_level: confloat(ge=0, le=1)
    confidence_score: confloat(ge=0, le=1)

class XSum_Hallucination_Detection:
    def __init__(self):
        # self.GPT_MODEL = "TURN OFF"
        # self.GPT_MODEL = "gpt-4-1106-preview"
        self.GPT_MODEL = "gpt-3.5-turbo-16k"
        self.tools = self.get_tool_list()
        self.system_result = defaultdict(list)
        self.test_hal = True

    def get_tool_list(self):
        schema = HallucinationDetection.schema()
        tools = [
            {
                "type": "function",
                "function": {
                    "name": "HallucinationDetection",
                    "description": "Detect the hallucination type, level and confidence of determining the hallucination in the summary of a document. ",
                    "parameters": schema
                }
            }
        ]
        return tools

    def parse_resp(self, document, summary):
        # print(f"Document: {document}")
        # print(f"Summary: {summary}")
        messages = [
            {
                "role": "system",
                "content": f"Your task is to determine if there is any hallucinations in the given summarization with the given context.  Do no make assumptions. Do not summarize the context, only response in JSON with hallucination_type: Literal['intrinsic', 'extrinsic', 'both', 'NULL'] hallucination_level: confloat(ge=0, le=1) confidence_score: confloat(ge=0, le=1) '."

            },{
                "role": "user",
                "content": f"Context: {document}, Summary: {summary}. Is the summary supported by the context?"
                # Poor: "content": f"Context: {document}, Summary: {summary}. Is the summary supported by the context? Let's work this out in a step by step way to be sure we have the right answer."
            }
        ]
        chat_response = self.chat_completion_request(messages, tools=self.tools)
        try:
            assistant_message = chat_response.json()["choices"][0]["message"]
            # pprint(assistant_message)
        except Exception as e:
            print("Unable to parse ChatCompletion response")
            print(f"Error: {chat_response.json()}")
            return e
        arguments_dict = json.loads(assistant_message['tool_calls'][0]['function']['arguments'])
        return arguments_dict

    @retry(wait=wait_random_exponential(multiplier=1, max=40), stop=stop_after_attempt(3))
    def chat_completion_request(self, messages, tools=None, tool_choice='auto'):
        headers = {
            "Content-Type": "application/json",
            "Authorization": "Bearer " + "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",
        }
        json_data = {"model": self.GPT_MODEL, "messages": messages, "seed": 123, "temperature": 0.2 }#,  "response_format": {"type": "json_object"}}
        if tools is not None:
            json_data.update({"tools": tools, "tool_choice":tool_choice})
            
        try:
            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=json_data,
            )
            return response
        except Exception as e:
            print("Unable to generate ChatCompletion response")
            print(f"Exception: {e}")
            return e

    def ask_gpt(self):
        # Load data from Excel
        df = pd.read_excel('./xsum-hal-system_statistics.xlsx', sheet_name=None, keep_default_na=False)

        npy_path = "./xsum_none_hal.npy" ## The pre-saved none hallucination entities in xsum dataset
        if self.test_hal:
            id_system_arr = np.load(npy_path)
            id_system_set = set()
            for (bbcid, system) in id_system_arr:
                id_system_set.add((bbcid, system))
            print(f"NULL summaries set count: {len(id_system_set)}")

        for system, data in df.items():
            # Initialize a list to hold all entries for this system.
            self.system_result[system] = []
            if not self.test_hal:
                if system != 'TranS2S': continue
            for _, row in tqdm(data.iterrows(), desc = f"{system}", total=len(data)):
                document_id = row['document_id']
                document = row['document']
                summary = row['summary']
                gt = str(row['hallucination_type'])

                if self.test_hal:
                    if (str(document_id), system) not in id_system_set:
                        continue
                    
                start = time.time()
                try:
                    arguments_dict = self.parse_resp(document, summary)
                    # arguments_dict = {'hallucination_type': 'ERROR', 'hallucination_level': -1, 'confidence_score': -1}
                except:
                    print("** ERROR ** error while communicating with GPT-3 API")
                    arguments_dict = {'hallucination_type': 'ERROR', 'hallucination_level': -1, 'confidence_score': -1}
                end = time.time()

                time_taken = end - start
                try:
                    result = {
                        'document_id': document_id,
                        'labeled_hallucination_type': gt,
                        'hallucination_type': arguments_dict['hallucination_type'],
                        'hallucination_level': arguments_dict['hallucination_level'],
                        'confidence_score': arguments_dict['confidence_score'],
                        'prediction': True if arguments_dict['hallucination_type'] != 'NULL' else False,
                        'label': True if gt != 'NULL' else False,
                        'time_taken': round(time_taken, 4),
                        'gpt_model': self.GPT_MODEL
                    }
                    # print(result)
                    # return
                except:
                    print("** ERROR ** key error in arguments_dict")
                    arguments_dict = {'hallucination_type': 'ERROR', 'hallucination_level': -1, 'confidence_score': -1}
                    result = {
                        'document_id': document_id,
                        'labeled_hallucination_type': gt,
                        'hallucination_type': arguments_dict['hallucination_type'],
                        'hallucination_level': arguments_dict['hallucination_level'],
                        'confidence_score': arguments_dict['confidence_score'],
                        'prediction': True if arguments_dict['hallucination_type'] != 'NULL' else False,
                        'label': True if gt != 'NULL' else False,
                        'time_taken': round(time_taken, 4),
                        'gpt_model': self.GPT_MODEL
                    }
                # Append the result for this document to the list for the system.
                self.system_result[system].append(result)
                # cnt+=1
                # if cnt == 30:
                #     self.save_statistics_to_excel()

        # Save the result to excel
        self.save_statistics_to_excel()

    # def save_statistics_to_excel(self, file_name='gpt3_predict_hallucination_results.xlsx'):
    #     with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    #         for system, results in self.system_result.items():
    #             # Convert the list of dictionaries to a DataFrame and write to a sheet.
    #             df = pd.DataFrame(results)
    #             df.to_excel(writer, sheet_name=system, index=False)
    #     print(f"Results have been saved to {file_name}")
    def save_statistics_to_excel(self, file_name='gpt3_predict_hallucination_results(gpt-3.5-turbo-0613).xlsx'):
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = file_name.replace('.xlsx', f'_{current_time}.xlsx')

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            for system, results in self.system_result.items():
                # Convert the list of dictionaries to a DataFrame.
                df = pd.DataFrame(results)

                # Save DataFrame to an Excel sheet.
                df.to_excel(writer, sheet_name=system, index=False)

                # Apply conditional formatting using openpyxl.
                workbook  = writer.book
                worksheet = writer.sheets[system]

                # Define the light red fill color.
                red_fill = PatternFill(start_color='FF9999',
                                       end_color='FF9999',
                                       fill_type='solid')

                for row in worksheet.iter_rows(min_row=2, max_col=len(df.columns), max_row=len(df) + 1):
                    if row[2].value == 'ERROR':  # Adjust the index as per your 'hallucination_type' column.
                        for cell in row:
                            cell.fill = red_fill

        print(f"Results have been saved to {file_name}")

# Example usage:
detector = XSum_Hallucination_Detection()
detector.ask_gpt()
